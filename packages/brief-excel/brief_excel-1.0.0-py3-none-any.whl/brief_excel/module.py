from openpyxl import Workbook
import os
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
from PIL import Image as PILImage
import io

def fpn_check(fpn: str) -> None:
    try:
        filename = './data/' + fpn + 'temp.xlsx'
        Workbook().save(filename)

    except TypeError:
        print("ERROR : file name is None!!! ---fpn_check()")
        
    except FileNotFoundError:
        dir_name = ''

        for char in fpn:
            if char == '/':
                break
            else:
                dir_name = dir_name + char
                
        if not os.path.exists('./data/' + dir_name.strip()):
            os.mkdir('./data/' + dir_name)
        
    except OSError:
        sanitized_name = ''

        for char in fpn:
            if char == '*':
                sanitized_name = sanitized_name + 'x' 
            else:
                sanitized_name = sanitized_name + char
        fpn = str(sanitized_name)

    finally:
        if (os.path.exists(filename)):
            os.remove(filename)

class BriefExcel:
    def __init__(self, name: str = None, pic_paths: list = None):
        """初始化 BriefExcel 实例
        
        Args:
            name: Excel 文件名
            pic_paths: 图片路径列表
        """
        self.name = name
        self.pic_paths = pic_paths if pic_paths else []
        
        self.wb = None
        self.ws = None
        self.ws_num = None
        self.cell = None
        self.pending_data: list = []
    
    def set_name(self, name: str) -> 'BriefExcel':
        """设置文件名
        
        Args:
            name: Excel 文件名（不含扩展名）
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        self.name = name
        return self
    
    def set_pic_paths(self, pic_paths: list) -> 'BriefExcel':
        """设置图片路径
        
        Args:
            pic_paths: 图片路径列表
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        self.pic_paths = pic_paths
        return self
       
    def create(self, fpn: str = None, row: int = 15, col: int = 15) -> 'BriefExcel':
        """创建新的 Excel 文件
        
        Args:
            fpn: Excel 文件名（不含扩展名）
            row: 初始单元格行号（默认15）
            col: 初始单元格列号（默认15）
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        if fpn is None:
            fpn = self.name
        
        fpn_check(fpn)
        fpn = 'data/'+ fpn +'.xlsx'

        if os.path.exists(fpn):
             os.remove(fpn)
        self.wb = Workbook()
        self.ws = self.wb.active
        self.cell = self.ws.cell(row, col)
        self.wb.save(fpn)

        self.ws_num = len(self.wb.sheetnames)
        return self

    def load(self, fpn: str = None, row: int = 1, col: int = 1) -> 'BriefExcel':
        """加载现有的 Excel 文件
        
        Args:
            fpn: Excel 文件名（不含扩展名）
            row: 初始单元格行号（默认1）
            col: 初始单元格列号（默认1）
            
        Returns:
            self: 返回当前实例以支持链式调用
            
        Raises:
            FileNotFoundError: 当指定的文件不存在时
            Exception: 当文件加载失败时
        """
        if fpn is None:
            fpn = self.name
            if fpn is None:
                raise ValueError("文件名不能为None")
        
        fpn_check(fpn)
        file_path = 'data/' + fpn + '.xlsx'

        # 检查文件是否存在
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel文件不存在: {file_path}")
        
        try:
            # 使用 openpyxl 加载现有文件
            from openpyxl import load_workbook
            self.wb = load_workbook(file_path)
            self.ws = self.wb.active
            self.cell = self.ws.cell(row, col)
            self.ws_num = len(self.wb.sheetnames)
            
            print(f"✓ 成功加载文件: {file_path}")
            return self
            
        except FileNotFoundError as e:
            print(f"✗ 文件加载失败: 文件不存在 - {e}")
            raise
        except PermissionError as e:
            print(f"✗ 文件加载失败: 权限不足 - {e}")
            raise
        except ValueError as e:
            print(f"✗ 文件加载失败: 文件格式错误 - {e}")
            raise
        except Exception as e:
            print(f"✗ 文件加载失败: 未知错误 - {e}")
            raise
    
    def ws_append(self, name: str = 'Sheet', num: int = 1) -> 'BriefExcel':
        """添加工作表
        
        Args:
            name: 工作表名称（默认'Sheet'）
            num: 添加的工作表数量（默认1）
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        for i in range(1, num + 1):
            self.wb.create_sheet(name, self.ws_num + i)
        self.ws_num = len(self.wb.sheetnames)
        return self
    
    def ws_location(self, title_or_num: str | int) -> 'BriefExcel':
        """定位到指定工作表
        
        Args:
            title_or_num: 工作表名称或索引
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        if isinstance(title_or_num, str):
            self.ws = self.wb[title_or_num]
        elif isinstance(title_or_num, int):
            self.ws = self.wb[self.wb.sheetnames[title_or_num]]
        return self

    def ws_move(self, num: int) -> 'BriefExcel':
        """移动当前工作表位置
        
        Args:
            num: 移动到的位置索引
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        self.wb.move_sheet(self.ws, num)
        return self
    
    def ws_copy(self, title_or_num: str | int) -> 'BriefExcel':
        """复制工作表
        
        Args:
            title_or_num: 要复制的工作表名称或索引
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        if isinstance(title_or_num, str):
            ws = self.wb[title_or_num]
        elif isinstance(title_or_num, int):
            ws = self.wb[self.wb.sheetnames[title_or_num]]

        self.ws = self.wb.copy_worksheet(ws)

        return self

    def ws_delete(self) -> 'BriefExcel':
        """删除当前工作表
        
        Returns:
            self: 返回当前实例以支持链式调用
        """
        del self.ws
        try:
            self.ws = self.wb[self.wb.sheetnames[0]]
        except IndexError:
            # 如果没有工作表了，设置为None
            self.ws = None
        return self
    
    def cell_location(self, row: int, col: int = 0) -> 'BriefExcel':
        """定位到指定单元格
        
        Args:
            row: 单元格行号
            col: 单元格列号（默认0，表示整行）
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        if col:
            self.cell = self.ws.cell(row, col)
        else:
            self.cell = self.ws[row]

        self.pending_data = []
        self.pending_data.append(self.cell.value)
        return self
    
    def cell_write(self, data: str) -> 'BriefExcel':
        """写入单个单元格数据
        
        Args:
            data: 要写入的字符串数据
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        self.cell.value = data
        return self
    
    def cell_row_write(self, row: int, col: int = 0, datas: list = None) -> 'BriefExcel':
        """写入整行数据
        
        Args:
            row: 行号
            col: 起始列号（默认0，表示从第一列开始）
            datas: 要写入的数据列表（如果为None，则使用pending_data）
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        if datas is None:
            datas = self.pending_data

        self.cell_location(row, col)

        cell_row = list(self.ws[self.cell.row])
        cell_row = cell_row[(self.cell.column - 1):]

        i = 0
        for data_item in datas:
            self.cell = cell_row[i]
            i = i + 1
            self.cell_write(data_item)
        self.wb.save('data/' + self.name + '.xlsx')
        return self

    def set_cell_style(self, row: int, col: int, 
                      font_name: str = None, font_size: int = None, font_bold: bool = False,
                      bg_color: str = None, font_color: str = None,
                      border: bool = False, align: str = None) -> 'BriefExcel':
        """设置单元格样式
        
        Args:
            row: 行号
            col: 列号
            font_name: 字体名称
            font_size: 字体大小
            font_bold: 是否加粗
            bg_color: 背景颜色（十六进制格式，如 'FF0000'）
            font_color: 字体颜色（十六进制格式，如 'FFFFFF'）
            border: 是否添加边框
            align: 对齐方式（'left', 'center', 'right', 'top', 'middle', 'bottom'）
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        self.cell_location(row, col)
        
        # 设置字体
        font = Font()
        if font_name:
            font.name = font_name
        if font_size:
            font.size = font_size
        if font_bold:
            font.bold = font_bold
        if font_color:
            font.color = font_color
        
        self.cell.font = font
        
        # 设置背景颜色
        if bg_color:
            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
            self.cell.fill = fill
        
        # 设置边框
        if border:
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            self.cell.border = thin_border
        
        # 设置对齐方式
        if align:
            alignment = Alignment()
            if align in ['left', 'center', 'right']:
                alignment.horizontal = align
            if align in ['top', 'middle', 'bottom']:
                alignment.vertical = align
            self.cell.alignment = alignment
        
        self.wb.save('data/' + self.name + '.xlsx')
        return self

    def set_column_width(self, col: int, width: float) -> 'BriefExcel':
        """设置列宽
        
        Args:
            col: 列号
            width: 列宽（字符数）
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        column_letter = get_column_letter(col)
        self.ws.column_dimensions[column_letter].width = width
        self.wb.save('data/' + self.name + '.xlsx')
        return self

    def set_row_height(self, row: int, height: float) -> 'BriefExcel':
        """设置行高
        
        Args:
            row: 行号
            height: 行高（磅值）
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        self.ws.row_dimensions[row].height = height
        self.wb.save('data/' + self.name + '.xlsx')
        return self

    def set_formula(self, row: int, col: int, formula: str) -> 'BriefExcel':
        """设置单元格公式
        
        Args:
            row: 行号
            col: 列号
            formula: Excel公式（如 '=SUM(A1:A10)'）
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        self.cell_location(row, col)
        self.cell.value = formula
        self.wb.save('data/' + self.name + '.xlsx')
        return self

    def apply_conditional_formatting(self, range_str: str, formula: str, 
                                   bg_color: str = None, font_color: str = None) -> 'BriefExcel':
        """应用条件格式
        
        Args:
            range_str: 单元格范围（如 'A1:B10'）
            formula: 条件公式（如 '=A1>100'）
            bg_color: 满足条件时的背景颜色
            font_color: 满足条件时的字体颜色
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        try:
            from openpyxl.formatting.rule import FormulaRule
            from openpyxl.styles import PatternFill, Font
            
            # 创建格式规则
            rule = FormulaRule(formula=[formula])
            
            # 设置格式
            if bg_color or font_color:
                fmt = {}
                if bg_color:
                    fmt['fill'] = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                if font_color:
                    fmt['font'] = Font(color=font_color)
                rule.format = fmt
            
            # 应用条件格式
            self.ws.conditional_formatting.add(range_str, rule)
            self.wb.save('data/' + self.name + '.xlsx')
            
        except ImportError:
            print("条件格式功能需要 openpyxl 完整支持")
        
        return self

    def merge_cells(self, range_str: str) -> 'BriefExcel':
        """合并单元格
        
        Args:
            range_str: 单元格范围（如 'A1:B2'）
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        self.ws.merge_cells(range_str)
        self.wb.save('data/' + self.name + '.xlsx')
        return self

    def freeze_panes(self, row: int = None, col: int = None) -> 'BriefExcel':
        """冻结窗格
        
        Args:
            row: 冻结的行号（从该行开始冻结）
            col: 冻结的列号（从该列开始冻结）
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        if row and col:
            self.ws.freeze_panes = self.ws.cell(row, col)
        elif row:
            self.ws.freeze_panes = self.ws.cell(row, 1)
        elif col:
            self.ws.freeze_panes = self.ws.cell(1, col)
        
        self.wb.save('data/' + self.name + '.xlsx')
        return self

    def set_number_format(self, row: int, col: int, format_str: str) -> 'BriefExcel':
        """设置数字格式
        
        Args:
            row: 行号
            col: 列号
            format_str: 数字格式字符串（如 '0.00', '¥#,##0.00', 'yyyy-mm-dd'）
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        self.cell_location(row, col)
        self.cell.number_format = format_str
        self.wb.save('data/' + self.name + '.xlsx')
        return self

    def _get_image_size(self, path: str) -> tuple:
        """获取图片的实际尺寸
        
        Args:
            path: 图片路径
            
        Returns:
            tuple: (width, height) 图片的实际宽高
        """
        try:
            with PILImage.open(path) as img:
                return img.size
        except Exception as e:
            print(f"获取图片尺寸失败: {e}")
            return (720, 1280)  # 默认尺寸

    def _resize_image(self, path: str, max_width: int = None, max_height: int = None) -> Image:
        """调整图片尺寸并转换为openpyxl Image对象
        
        Args:
            path: 图片路径
            max_width: 最大宽度（可选）
            max_height: 最大高度（可选）
            
        Returns:
            Image: 调整后的openpyxl Image对象
        """
        try:
            # 打开原始图片
            pil_img = PILImage.open(path)
            original_width, original_height = pil_img.size
            
            # 计算调整后的尺寸
            if max_width and max_height:
                # 保持宽高比缩放
                ratio = min(max_width / original_width, max_height / original_height)
                new_width = int(original_width * ratio)
                new_height = int(original_height * ratio)
            elif max_width:
                # 只限制宽度
                ratio = max_width / original_width
                new_width = max_width
                new_height = int(original_height * ratio)
            elif max_height:
                # 只限制高度
                ratio = max_height / original_height
                new_width = int(original_width * ratio)
                new_height = max_height
            else:
                # 不调整尺寸
                new_width, new_height = original_width, original_height
            
            # 如果不需要调整尺寸，直接返回原始图片
            if (new_width, new_height) == (original_width, original_height):
                return Image(path)
            
            # 调整图片尺寸
            pil_img = pil_img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
            
            # 保存到临时文件
            import tempfile
            temp_file = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            pil_img.save(temp_file, format='PNG')
            temp_file.close()
            
            # 创建openpyxl Image对象
            img = Image(temp_file.name)
            img.width, img.height = new_width, new_height
            
            # 清理临时文件
            os.unlink(temp_file.name)
            
            return img
            
        except Exception as e:
            print(f"图片处理失败: {e}")
            # 失败时返回原始图片
            return Image(path)

    def insert_pic_advanced(self, row: int, col: int = 0, path: str = None, 
                          max_width: int = None, max_height: int = None,
                          scale: float = 1.0, keep_aspect_ratio: bool = True) -> 'BriefExcel':
        """高级图片插入功能
        
        Args:
            row: 行号
            col: 列号（默认0，表示第一列）
            path: 图片路径
            max_width: 最大宽度（可选）
            max_height: 最大高度（可选）
            scale: 缩放比例（默认1.0）
            keep_aspect_ratio: 是否保持宽高比（默认True）
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        if path is None:
            if self.pic_paths:
                path = self.pic_paths[0]
            else:
                raise ValueError("没有提供图片路径")
        
        # 检查图片文件是否存在
        if not os.path.exists(path):
            raise FileNotFoundError(f"图片文件不存在: {path}")
        
        self.cell_location(row, col)
        
        # 处理图片
        img = self._resize_image(path, max_width, max_height)
        
        # 应用缩放比例
        if scale != 1.0:
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)
        
        # 调整单元格尺寸以适应图片
        column_letter = get_column_letter(self.cell.column)
        self.ws.column_dimensions[column_letter].width = max(img.width * 0.14, 8.43)  # 最小宽度
        self.ws.row_dimensions[row].height = max(img.height * 0.78, 15)  # 最小高度
        
        # 插入图片
        self.ws.add_image(img, self.cell.coordinate)
        self.wb.save('data/' + self.name + '.xlsx')
        
        return self

    def insert_pic_from_url(self, row: int, col: int = 0, url: str = None, 
                          max_width: int = None, max_height: int = None) -> 'BriefExcel':
        """从URL插入图片
        
        Args:
            row: 行号
            col: 列号（默认0，表示第一列）
            url: 图片URL
            max_width: 最大宽度（可选）
            max_height: 最大高度（可选）
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        if url is None:
            raise ValueError("没有提供图片URL")
        
        try:
            import requests
            from io import BytesIO
            
            # 下载图片
            response = requests.get(url)
            response.raise_for_status()
            
            # 创建临时文件
            temp_path = f"temp_image_{hash(url)}.png"
            with open(temp_path, 'wb') as f:
                f.write(response.content)
            
            # 使用高级插入方法
            self.insert_pic_advanced(row, col, temp_path, max_width, max_height)
            
            # 清理临时文件
            os.remove(temp_path)
            
            return self
            
        except ImportError:
            print("请安装requests库: pip install requests")
            raise
        except Exception as e:
            print(f"从URL插入图片失败: {e}")
            raise
    
    def cell_row_read(self, row: int) -> 'BriefExcel':
        """读取整行数据
        
        Args:
            row: 要读取的行号
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        self.pending_data = []
        try:
            # 获取指定行的所有单元格
            for col in range(1, self.ws.max_column + 1):
                cell = self.ws.cell(row=row, column=col)
                self.pending_data.append(cell.value)
        except IndexError as e:
            print(f"读取行数据错误: 行索引越界 - {e}")
            self.pending_data = []
        except ValueError as e:
            print(f"读取行数据错误: 无效的行号 - {e}")
            self.pending_data = []
        except Exception as e:
            print(f"读取行数据错误: 未知错误 - {e}")
            self.pending_data = []
        return self
       
    def cell_col_write(self, row: int, col: int = 0, datas: list = None) -> 'BriefExcel':
        """写入整列数据
        
        Args:
            row: 起始行号
            col: 列号
            datas: 要写入的数据列表（如果为None，则使用pending_data）
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        if datas is None:
            datas = self.pending_data

        self.cell_location(row, col)

        cell_col = list(self.ws[self.cell.coordinate[0]])
        cell_col = cell_col[(self.cell.row - 1):]

        i = 0
        for data_item in datas:
            self.cell = cell_col[i]
            i = i + 1
            self.cell_write(data_item)
        self.wb.save('data/' + self.name + '.xlsx')
        return self
    
    def cell_col_read(self, col: int) -> 'BriefExcel':
        """读取整列数据
        
        Args:
            col: 要读取的列号
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        self.pending_data = []
        try:
            # 获取指定列的所有单元格
            for row in range(1, self.ws.max_row + 1):
                cell = self.ws.cell(row=row, column=col)
                self.pending_data.append(cell.value)
        except IndexError as e:
            print(f"读取列数据错误: 列索引越界 - {e}")
            self.pending_data = []
        except ValueError as e:
            print(f"读取列数据错误: 无效的列号 - {e}")
            self.pending_data = []
        except Exception as e:
            print(f"读取列数据错误: 未知错误 - {e}")
            self.pending_data = []
        return self
    
    def cell_area_write(self, row: int, col: int = 1, length: int = 2, width: int = 2, datas: list = None) -> 'BriefExcel':
        """写入区域数据
        
        Args:
            row: 起始行号
            col: 起始列号（默认1）
            length: 区域长度（列数，默认2）
            width: 区域宽度（行数，默认2）
            datas: 要写入的数据列表（如果为None，则使用pending_data）
            
        Returns:
            self: 返回当前实例以支持链式调用
        """
        if datas is None:
            datas = self.pending_data

        data_i = 0
        
        try:
            # 确保数据是扁平化的一维列表
            flat_data = []
            for item in datas:
                if isinstance(item, list):
                    flat_data.extend(item)
                else:
                    flat_data.append(item)
            
            for j in range(width):
                for i in range(length):
                    if data_i < len(flat_data):
                        cell = self.ws.cell(row=row + j, column=col + i)
                        # 确保数据是字符串类型
                        cell_value = str(flat_data[data_i]) if flat_data[data_i] is not None else ""
                        cell.value = cell_value
                        data_i += 1
            # 保存文件
            self.wb.save('data/' + self.name + '.xlsx')
        except IndexError as e:
            print(f"区域写入错误: 区域索引越界 - {e}")
        except ValueError as e:
            print(f"区域写入错误: 无效的区域参数 - {e}")
        except Exception as e:
            print(f"区域写入错误: 未知错误 - {e}")
        
        return self
    
    def cell_area_read(self, row: int, col: int = 1, length: int = 2, width: int = 2) -> 'BriefExcel':
        """读取区域数据
        
        Args:
            row: 起始行号
            col: 起始列号（默认1）
            length: 区域长度（列数，默认2）
            width: 区域宽度（行数，默认2）
            
        Returns:
            self: 返回当前实例以支持链式调用，读取的数据存储在 pending_data 中
        """
        self.pending_data = []
        try:
            for j in range(width):
                for i in range(length):
                    cell = self.ws.cell(row=row + j, column=col + i)
                    self.pending_data.append(cell.value)
        except IndexError as e:
            print(f"区域读取错误: 区域索引越界 - {e}")
            self.pending_data = []
        except ValueError as e:
            print(f"区域读取错误: 无效的区域参数 - {e}")
            self.pending_data = []
        except Exception as e:
            print(f"区域读取错误: 未知错误 - {e}")
            self.pending_data = []
        return self
    
    def insert_pic(self, row: int, col: int = 0, path: str = None, 
                  width: int = 720, height: int = 1280,
                  max_width: int = None, max_height: int = None,
                  scale: float = 1.0) -> 'BriefExcel':
        """插入单张图片
        
        Args:
            row: 行号
            col: 列号（默认0，表示第一列）
            path: 图片路径
            width: 图片宽度（默认720）
            height: 图片高度（默认1280）
            max_width: 最大宽度限制（可选）
            max_height: 最大高度限制（可选）
            scale: 缩放比例（默认1.0）
            
        Returns:
            self: 返回当前实例以支持链式调用
            
        Raises:
            FileNotFoundError: 当图片文件不存在时
            ValueError: 当没有提供图片路径时
        """
        if path is None:
            if self.pic_paths:
                path = self.pic_paths[0]
            else:
                raise ValueError("没有提供图片路径")
        
        # 检查图片文件是否存在
        if not os.path.exists(path):
            raise FileNotFoundError(f"图片文件不存在: {path}")
        
        self.cell_location(row, col)

        # 获取图片实际尺寸
        try:
            with PILImage.open(path) as pil_img:
                original_width, original_height = pil_img.size
                
                # 计算最终尺寸
                if max_width and max_height:
                    # 保持宽高比缩放
                    ratio = min(max_width / original_width, max_height / original_height)
                    final_width = int(original_width * ratio)
                    final_height = int(original_height * ratio)
                elif max_width:
                    # 只限制宽度
                    ratio = max_width / original_width
                    final_width = max_width
                    final_height = int(original_height * ratio)
                elif max_height:
                    # 只限制高度
                    ratio = max_height / original_height
                    final_width = int(original_width * ratio)
                    final_height = max_height
                else:
                    # 使用默认尺寸或原始尺寸
                    final_width = width if width else original_width
                    final_height = height if height else original_height
                
                # 应用缩放比例
                final_width = int(final_width * scale)
                final_height = int(final_height * scale)
                
        except Exception as e:
            print(f"获取图片尺寸失败，使用默认尺寸: {e}")
            final_width, final_height = width, height

        imgsize = (final_width / 4, final_height / 4)
        self.ws.column_dimensions[get_column_letter(self.cell.column)].width = max(imgsize[0] * 0.14, 8.43)
        
        img = Image(path)
        img.width, img.height = imgsize

        self.ws.add_image(img, self.cell.coordinate)
        self.ws.row_dimensions[row].height = max(imgsize[1] * 0.78, 15)
        self.wb.save('data/' + self.name + '.xlsx')
        return self
    
    def insert_pic_row(self, row: int, col: int = 0, paths: list = None, width: int = 720, height: int = 1280) -> 'BriefExcel':
        """在行中插入多张图片
        
        Args:
            row: 起始行号
            col: 起始列号（默认0，表示从第一列开始）
            paths: 图片路径列表（如果为None，则使用实例的pic_paths）
            width: 图片宽度（默认720）
            height: 图片高度（默认1280）
            
        Returns:
            self: 返回当前实例以支持链式调用
            
        Raises:
            FileNotFoundError: 当图片文件不存在时
            ValueError: 当没有提供图片路径时
        """
        if paths is None:
            paths = self.pic_paths
        
        if not paths:
            raise ValueError("没有提供图片路径")
        
        self.cell_location(row, col)

        # 获取当前行的单元格
        cell_row = list(self.ws[self.cell.row])
        cell_row = cell_row[(self.cell.column - 1):]

        imgsize = (width / 4, height / 4)
        self.ws.column_dimensions[get_column_letter(self.cell.column)].width = imgsize[0] * 0.14
        
        i = 0
        for image_path in paths:
            if i >= len(cell_row):
                break  # 防止索引越界
                
            self.cell = cell_row[i]
            
            # 检查图片文件是否存在
            if not os.path.exists(image_path):
                raise FileNotFoundError(f"图片文件不存在: {image_path}")
            
            img = Image(image_path)
            img.width, img.height = imgsize
            self.ws.add_image(img, self.cell.coordinate)
            self.ws.row_dimensions[row].height = imgsize[1] * 0.78
            i = i + 1

        self.wb.save('data/' + self.name + '.xlsx')
        return self

    def insert_pic_col(self, row: int, col: int = 0, paths: list = None, width: int = 720, height: int = 1280) -> 'BriefExcel':
        """在列中插入多张图片
        
        Args:
            row: 起始行号
            col: 列号
            paths: 图片路径列表（如果为None，则使用实例的pic_paths）
            width: 图片宽度（默认720）
            height: 图片高度（默认1280）
            
        Returns:
            self: 返回当前实例以支持链式调用
            
        Raises:
            FileNotFoundError: 当图片文件不存在时
            ValueError: 当没有提供图片路径时
        """
        if paths is None:
            paths = self.pic_paths
        
        if not paths:
            raise ValueError("没有提供图片路径")
        
        self.cell_location(row, col)

        # 获取当前列的单元格
        column_letter = get_column_letter(col if col > 0 else self.cell.column)
        cell_col = list(self.ws[column_letter])
        cell_col = cell_col[(self.cell.row - 1):]

        imgsize = (width / 4, height / 4)
        self.ws.column_dimensions[column_letter].width = imgsize[0] * 0.14
        
        i = 0
        for image_path in paths:
            if i >= len(cell_col):
                break  # 防止索引越界
                
            self.cell = cell_col[i]
            
            # 检查图片文件是否存在
            if not os.path.exists(image_path):
                raise FileNotFoundError(f"图片文件不存在: {image_path}")
            
            img = Image(image_path)
            img.width, img.height = imgsize
            self.ws.add_image(img, self.cell.coordinate)
            self.ws.row_dimensions[self.cell.row].height = imgsize[1] * 0.78
            i = i + 1
            
        self.wb.save('data/' + self.name + '.xlsx')
        return self
