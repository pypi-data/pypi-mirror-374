"""Entropy analysis CLI commands with streaming Excel export."""

from __future__ import annotations

import time
from pathlib import Path
from typing import TYPE_CHECKING, NamedTuple

import rich_click as click
from dependency_injector.wiring import Provide, inject

from kp_ssf_tools.analyze.models.types import EntropyLevel
from kp_ssf_tools.analyze.services.interfaces import CredentialScanOptions
from kp_ssf_tools.containers.application import ApplicationContainer
from kp_ssf_tools.core.services.excel_export.streaming import StreamingExcelExporter

if TYPE_CHECKING:
    import pandas as pd

    from kp_ssf_tools.analyze.models.analysis import (
        CredentialAnalysisResult,
        CredentialPattern,
    )
    from kp_ssf_tools.analyze.models.configuration import AnalysisConfiguration
    from kp_ssf_tools.analyze.services.entropy.analyzer import EntropyAnalyzer
    from kp_ssf_tools.analyze.services.interfaces import CredentialDetectionProtocol
    from kp_ssf_tools.core.services.config.models import GlobalConfiguration
    from kp_ssf_tools.core.services.config.service import ConfigurationService
    from kp_ssf_tools.core.services.config.types import ConfigDict
    from kp_ssf_tools.core.services.excel_export.service import ExcelExportService
    from kp_ssf_tools.core.services.excel_export.sheet_management import (
        DefaultSheetNameSanitizer,
    )
    from kp_ssf_tools.core.services.file_processing.discovery import (
        FileDiscoveryService,
    )
    from kp_ssf_tools.core.services.rich_output.service import RichOutputService
    from kp_ssf_tools.core.services.timestamp.service import TimestampService


class AnalysisSummary(NamedTuple):
    """Summary data for analysis results."""

    total_files_analyzed: int
    total_files: int
    total_high_risk_regions: int
    total_time: float


class ProcessingConfig(NamedTuple):
    """Configuration for file processing."""

    file_block_size: int | None
    analysis_block_size: int | None
    step_size: int | None
    include_samples: bool


class ProcessingContext(NamedTuple):
    """Context for file processing with analyzer and configs."""

    analyzer: EntropyAnalyzer
    entropy_config: AnalysisConfiguration
    global_config: GlobalConfiguration
    rich_output: RichOutputService


class ExportContext(NamedTuple):
    """Context for Excel export operations."""

    excel_service: ExcelExportService
    timestamp_service: TimestampService
    rich_output: RichOutputService


def _build_cli_overrides(
    file_block_size: int | None,
    analysis_block_size: int | None,
    step_size: int | None,
) -> ConfigDict:
    """Build CLI overrides for entropy configuration."""
    cli_overrides: ConfigDict = {}
    if file_block_size is not None:
        analysis_section = cli_overrides.setdefault("analysis", {})
        if isinstance(analysis_section, dict):
            analysis_section["file_chunk_size"] = file_block_size
    if analysis_block_size is not None:
        analysis_section = cli_overrides.setdefault("analysis", {})
        if isinstance(analysis_section, dict):
            analysis_section["analysis_block_size"] = analysis_block_size
    if step_size is not None:
        analysis_section = cli_overrides.setdefault("analysis", {})
        if isinstance(analysis_section, dict):
            analysis_section["step_size"] = step_size
    return cli_overrides


def _discover_and_filter_files(
    target: Path,
    file_discovery: FileDiscoveryService,
    *,
    no_recurse: bool,
    ignore_pattern: tuple[str, ...],
    rich_output: RichOutputService,
) -> list[Path]:
    """Discover files to analyze and apply ignore patterns."""
    if target.is_file():
        rich_output.info(f"Analyzing single file: {target}")
        return [target]

    if not target.is_dir():
        rich_output.error(f"Target path is neither file nor directory: {target}")
        return []

    files = file_discovery.discover_files_by_pattern(
        base_path=target,
        pattern="*",  # All files
        recursive=not no_recurse,
    )

    # Filter out files matching any ignore pattern
    if ignore_pattern:
        import fnmatch

        filtered_files = []
        for f in files:
            if not any(
                fnmatch.fnmatch(str(f), pat) or pat in str(f) for pat in ignore_pattern
            ):
                filtered_files.append(f)
        files = filtered_files

    if not files:
        recurse_msg = "recursively" if not no_recurse else "in current directory only"
        rich_output.warning(f"No files found {recurse_msg}: {target}")
        return []

    recurse_msg = "recursively" if not no_recurse else "non-recursively"
    rich_output.info(f"Found {len(files)} files to analyze {recurse_msg}")
    return files


def _check_excel_limits(
    files_to_analyze: list[Path],
    risk_level: EntropyLevel,
    step_size_val: int,
    rich_output: RichOutputService,
) -> bool:
    """Check for potential Excel limit issues and warn user. Returns True if should abort."""
    excel_region_warning_threshold = 500_000
    total_size = sum(f.stat().st_size for f in files_to_analyze)
    estimated_regions = total_size // step_size_val

    if (
        risk_level <= EntropyLevel.LOW
        and estimated_regions > excel_region_warning_threshold
    ):
        rich_output.warning(
            f"Low risk threshold with ~{estimated_regions:,} estimated regions. "
            f"Excel has a {StreamingExcelExporter.MAX_ROWS:,} row limit. "
            "Consider using a higher threshold (--risk-threshold=medium_high)",
        )
        return not click.confirm("Continue anyway?")
    return False


def _format_file_display(file_path: Path, max_length: int = 50) -> str:
    """Format file path for display with truncation if needed."""
    if len(str(file_path)) > max_length:
        return f"...{str(file_path)[-(max_length - 3) :]}"
    return str(file_path)


def _process_files(
    files_to_analyze: list[Path],
    output_path: Path,
    risk_level: EntropyLevel,
    context: ProcessingContext,
    processing_config: ProcessingConfig,
) -> tuple[int, int, float]:
    """Process all files and return (files_analyzed, high_risk_regions, total_time)."""
    with StreamingExcelExporter(output_path, risk_level) as exporter:
        start_time = time.time()
        total_files_analyzed = 0
        total_high_risk_regions = 0

        for file_index, file_path in enumerate(files_to_analyze, 1):
            file_display = _format_file_display(file_path)
            context.rich_output.info(
                f"[{file_index}/{len(files_to_analyze)}] Analyzing: {file_display}",
            )

            try:
                # Stream analysis results directly to Excel
                total_regions, high_risk_regions = exporter.process_file_streaming(
                    file_path,
                    context.analyzer,
                    file_chunk_size=processing_config.file_block_size
                    or context.entropy_config.analysis.file_chunk_size,
                    analysis_block_size=processing_config.analysis_block_size
                    or context.entropy_config.analysis.analysis_block_size,
                    step_size=processing_config.step_size
                    or context.entropy_config.analysis.step_size,
                    include_samples=processing_config.include_samples or False,
                )

                total_files_analyzed += 1
                total_high_risk_regions += high_risk_regions

                if context.global_config.output.verbose:
                    context.rich_output.debug(
                        f"  Processed {total_regions:,} regions, "
                        f"found {high_risk_regions:,} high-risk regions",
                    )

            except Exception as e:  # noqa: BLE001
                # Broad exception catch justified: analysis may fail for any file due to I/O, format, or analyzer errors
                context.rich_output.error(f"  Failed to analyze {file_path.name}: {e}")
                if context.global_config.output.verbose:
                    import traceback

                    context.rich_output.error(traceback.format_exc())
                continue

        total_time = time.time() - start_time

        # Check if Excel limit warning should be shown
        if exporter.warned_about_limit:
            context.rich_output.warning(
                "Excel row limit was reached. Some regions may not be included. "
                "Consider using a higher --risk-threshold to reduce output.",
            )

        return total_files_analyzed, total_high_risk_regions, total_time


def _report_summary(
    output_path: Path,
    summary: AnalysisSummary,
    rich_output: RichOutputService,
) -> None:
    """Report final analysis summary."""
    rich_output.success(f"Analysis complete! Results saved to: {output_path}")
    rich_output.info(
        f"Summary:\n"
        f"  Files analyzed: {summary.total_files_analyzed}/{summary.total_files}\n"
        f"  High-risk regions found: {summary.total_high_risk_regions:,}\n"
        f"  Processing time: {summary.total_time:.2f} seconds\n"
        f"  Average time per file: {summary.total_time / max(1, summary.total_files_analyzed):.2f} seconds",
    )


@click.group(name="analyze")
def analyze_group() -> None:
    """Security analysis commands for PCI SSF 2.3 compliance."""


@analyze_group.command("entropy")
@click.argument("target", type=click.Path(exists=True, path_type=Path))
@click.option(
    "--ignore-pattern",
    multiple=True,
    help="Glob pattern(s) to ignore when searching for files (e.g. --ignore-pattern='__pycache__' --ignore-pattern='*.egg-info')",
)
@click.option(
    "--risk-threshold",
    type=click.Choice(
        ["very_low", "low", "medium", "medium_high", "high", "critical"],
        case_sensitive=False,
    ),
    default="medium_high",
    help="Minimum risk level for regions to include in analysis (default: medium_high)",
)
@click.option(
    "--file-block-size",
    type=int,
    help="File I/O block size in bytes (default: 65536)",
)
@click.option(
    "--analysis-block-size",
    type=int,
    help="Analysis block size in bytes (default: 64)",
)
@click.option(
    "--step-size",
    type=int,
    help="Step size for sliding window analysis (default: 16)",
)
@click.option(
    "--no-recurse",
    is_flag=True,
    help="Disable recursive directory analysis (analyze current directory only)",
)
@click.option(
    "--include-samples",
    is_flag=True,
    help="Include data samples in region analysis (increases file size)",
)
@inject
def entropy(  # noqa: PLR0913
    target: Path,
    risk_threshold: str,
    file_block_size: int | None,
    analysis_block_size: int | None,
    step_size: int | None,
    ignore_pattern: tuple[str, ...],
    *,
    no_recurse: bool,
    include_samples: bool,
    analyzer: EntropyAnalyzer = Provide[ApplicationContainer.entropy.analyzer],
    rich_output: RichOutputService = Provide[ApplicationContainer.core.rich_output],
    file_discovery: FileDiscoveryService = Provide[
        ApplicationContainer.core.file_discoverer
    ],
    global_config_service: ConfigurationService[GlobalConfiguration] = Provide[
        ApplicationContainer.core.global_config_service
    ],
    entropy_config_service: ConfigurationService[AnalysisConfiguration] = Provide[
        ApplicationContainer.core.entropy_config_service
    ],
    timestamp_service: TimestampService = Provide[ApplicationContainer.core.timestamp],
) -> None:
    """
    Analyze entropy of files for PCI SSF 2.3 compliance.

    Performs Shannon entropy analysis using content-aware thresholds to detect
    potentially suspicious patterns in files. Results are streamed directly to
    Excel with minimal memory usage.

    **Arguments:**
    ```
    TARGET                  Path to file or directory to analyze
    ```

    **Examples:**
    ```
    # Basic file analysis
    ssf_tools analyze entropy sample.bin

    # Analyze with higher risk threshold (fewer results)
    ssf_tools analyze entropy sample.bin --risk-threshold high

    # Analyze with custom block size
    ssf_tools analyze entropy sample.bin --analysis-block-size 128

    # Override file type detection
    ssf_tools analyze entropy app.exe --force-file-type windows_pe

    # Analyze directory non-recursively
    ssf_tools analyze entropy data/ --no-recurse
    ```
    """
    try:
        # Build configuration overrides and load configs
        cli_overrides = _build_cli_overrides(
            file_block_size,
            analysis_block_size,
            step_size,
        )
        global_config = global_config_service.load_config()
        entropy_config = entropy_config_service.load_config(
            command_overrides=cli_overrides,
        )

        if global_config.output.verbose:
            rich_output.debug("Verbose mode enabled")

        # Setup analysis parameters
        risk_level = EntropyLevel[risk_threshold.upper()]
        timestamp = timestamp_service.format_filename_now()
        output_path = Path(f"entropy-analysis-{timestamp}.xlsx")

        # Discover and filter files to analyze
        files_to_analyze = _discover_and_filter_files(
            target,
            file_discovery,
            no_recurse=no_recurse,
            ignore_pattern=ignore_pattern,
            rich_output=rich_output,
        )
        if not files_to_analyze:
            return

        # Check Excel limits and warn if necessary
        step_size_val = step_size or entropy_config.analysis.step_size
        if _check_excel_limits(
            files_to_analyze,
            risk_level,
            step_size_val,
            rich_output,
        ):
            return

        # Process files
        rich_output.info(
            f"Starting entropy analysis with {risk_level.value} risk threshold",
        )
        rich_output.info(f"Output will be saved to: {output_path}")

        processing_config = ProcessingConfig(
            file_block_size=file_block_size,
            analysis_block_size=analysis_block_size,
            step_size=step_size,
            include_samples=include_samples,
        )

        context = ProcessingContext(
            analyzer=analyzer,
            entropy_config=entropy_config,
            global_config=global_config,
            rich_output=rich_output,
        )

        total_files_analyzed, total_high_risk_regions, total_time = _process_files(
            files_to_analyze,
            output_path,
            risk_level,
            context,
            processing_config,
        )

        # Report final summary
        summary = AnalysisSummary(
            total_files_analyzed=total_files_analyzed,
            total_files=len(files_to_analyze),
            total_high_risk_regions=total_high_risk_regions,
            total_time=total_time,
        )
        _report_summary(output_path, summary, rich_output)

    except Exception as e:
        rich_output.error(f"Analysis failed: {e}")
        if "global_config" in locals() and global_config.output.verbose:
            import traceback

            rich_output.error(traceback.format_exc())
        raise


@analyze_group.command("credentials")
@click.argument("target", type=click.Path(exists=True, path_type=Path))
@click.option(
    "--recursive/--no-recursive",
    default=True,
    help="Search directories recursively for files to analyze",
)
@click.option(
    "--file-extensions",
    multiple=True,
    help="File extensions to include (e.g., .py .js .txt). If not specified, all text files are analyzed",
)
@click.option(
    "--context-lines",
    type=int,
    default=3,
    help="Number of context lines to show around matches",
)
@click.option(
    "--scan-binary/--no-scan-binary",
    default=True,
    help="Whether to scan binary files for embedded credentials",
)
@click.option(
    "--max-binary-size",
    type=int,
    default=10,
    help="Maximum size in MB for binary files to scan",
)
@inject
def credentials(  # noqa: PLR0913
    target: Path,
    *,
    recursive: bool,
    file_extensions: tuple[str, ...],
    context_lines: int,
    scan_binary: bool,
    max_binary_size: int,
    credential_service: CredentialDetectionProtocol = Provide[
        ApplicationContainer.analysis.active_credential_service
    ],
    rich_output: RichOutputService = Provide[ApplicationContainer.core.rich_output],
    excel_service: ExcelExportService = Provide[
        ApplicationContainer.core.excel_export_service
    ],
    timestamp_service: TimestampService = Provide[ApplicationContainer.core.timestamp],
    global_config_service: ConfigurationService[GlobalConfiguration] = Provide[
        ApplicationContainer.core.global_config_service
    ],
    analysis_config_service: ConfigurationService[AnalysisConfiguration] = Provide[
        ApplicationContainer.core.entropy_config_service
    ],
) -> None:
    """
    Detect credentials in files for PCI SSF 2.3 compliance.

    Analyzes files for embedded credentials including usernames, passwords,
    API keys, and other sensitive information. Uses wordlists from SecLists
    and regex patterns to identify potential security issues.

    Results are automatically exported to Excel with per-file worksheets
    and a summary sheet. Output filename: analyze-credentials-<timestamp>.xlsx

    **Arguments:**
    ```
    TARGET                  Path to file or directory to analyze
    ```

    **Examples:**
    ```
    # Basic credential detection
    ssf_tools analyze credentials sample.py

    # Analyze specific file types only
    ssf_tools analyze credentials data/ --file-extensions .py --file-extensions .js

    # Include more context around matches
    ssf_tools analyze credentials config/ --context-lines 5

    # Skip binary files to speed up analysis
    ssf_tools analyze credentials project/ --no-scan-binary
    ```
    """
    # Load configurations
    global_config = global_config_service.load_config()
    analysis_config = analysis_config_service.load_config()

    # Set verbose mode if requested
    if global_config.output.verbose:
        rich_output.debug("Verbose mode enabled")

    # Convert file extensions to list
    extensions_list = list(file_extensions) if file_extensions else None

    try:
        # Convert the Pydantic configuration to the dict format expected by the service
        config_dict = {
            "credentials": {
                "enabled": analysis_config.credentials.enabled,
                "cache_duration_hours": analysis_config.credentials.cache_duration_hours,
                "auto_download": analysis_config.credentials.auto_download,
                "wordlist_sources": analysis_config.credentials.wordlist_sources,
            },
        }

        # Perform credential analysis
        scan_options = CredentialScanOptions(
            recursive=recursive,
            file_extensions=tuple(extensions_list) if extensions_list else (),
            context_lines=context_lines,
            scan_binary_files=scan_binary,
            max_binary_size_mb=max_binary_size,
        )

        result = credential_service.analyze_files(
            target_paths=[target],
            config=config_dict,
            options=scan_options,
        )

        # Export to Excel if results found
        if result and result.patterns:
            export_context = ExportContext(
                excel_service=excel_service,
                timestamp_service=timestamp_service,
                rich_output=rich_output,
            )
            _export_credentials_to_excel(result, export_context)

        # Display results in console
        credential_display_limit = 10
        if result and result.patterns:
            rich_output.warning(f"Found {len(result.patterns)} potential credentials")
            for pattern in result.patterns[:credential_display_limit]:  # Show first N
                rich_output.info(
                    f"  {pattern.pattern_type}: {pattern.value[:50]}... "
                    f"(line {pattern.line_start})",
                )
            if len(result.patterns) > credential_display_limit:
                rich_output.info(
                    f"  ... and {len(result.patterns) - credential_display_limit} more",
                )
            rich_output.info("Complete results have been exported to Excel")
        else:
            rich_output.success("No credentials detected in analyzed files")

    except Exception as e:
        rich_output.error(f"Credential analysis failed: {e}")
        if global_config.output.verbose:
            import traceback

            rich_output.error(traceback.format_exc())
        raise


def _get_unique_worksheet_name(
    file_path: Path,
    used_names: set[str],
    sanitizer: DefaultSheetNameSanitizer,
) -> str:
    """Generate unique worksheet name with suffix if needed."""
    # Constants for worksheet naming
    max_base_name_length = 28
    max_suffix = 99

    base_name = sanitizer.sanitize_sheet_name(file_path.name)
    # Truncate to leave room for suffix
    if len(base_name) > max_base_name_length:
        base_name = base_name[:max_base_name_length]

    candidate = base_name
    suffix = 1
    while candidate in used_names:
        candidate = f"{base_name}_{suffix:02d}"
        suffix += 1
        if suffix > max_suffix:  # Safety limit
            break

    used_names.add(candidate)
    return candidate


def _build_summary_data(
    result: CredentialAnalysisResult,
    files_with_patterns: dict[Path, list[CredentialPattern]],
    worksheet_name_mapping: dict[str, str],
) -> list[dict[str, str | int]]:
    """Build summary data for all processed files."""
    summary_data: list[dict[str, str | int]] = []

    for file_path in result.processed_files:
        patterns = files_with_patterns.get(file_path, [])
        pattern_counts: dict[str, int] = {}
        risk_counts: dict[str, int] = {}

        for pattern in patterns:
            pattern_type = str(pattern.pattern_type)
            risk_level = str(pattern.risk_level)
            pattern_counts[pattern_type] = pattern_counts.get(pattern_type, 0) + 1
            risk_counts[risk_level] = risk_counts.get(risk_level, 0) + 1

        # Get worksheet name for files with patterns
        worksheet_name = worksheet_name_mapping.get(str(file_path), "")

        summary_data.append(
            {
                "File_Path": str(file_path),
                "Worksheet_Name": worksheet_name,
                "Total_Patterns": len(patterns),
                "Pattern_Types": ", ".join(pattern_counts.keys())
                if pattern_counts
                else "None",
                "Critical_Risk_Count": int(risk_counts.get("critical", 0)),
                "High_Risk_Count": int(risk_counts.get("high", 0)),
                "Medium_Risk_Count": int(risk_counts.get("medium", 0)),
                "Low_Risk_Count": int(risk_counts.get("low", 0)),
                "Info_Risk_Count": int(risk_counts.get("info", 0)),
            },
        )

    return summary_data


def _add_hyperlinks_to_summary(
    writer: pd.ExcelWriter,
    summary_df: pd.DataFrame,
) -> None:
    """Add hyperlinks to summary worksheet for files with findings."""
    if not (hasattr(writer, "book") and hasattr(writer, "sheets")):
        return

    summary_sheet = writer.sheets["Summary"]

    # Add hyperlinks in the Worksheet_Name column (column B, index 1)
    for row_idx, (_, row) in enumerate(
        summary_df.iterrows(),
        start=2,
    ):  # start=2 for 1-indexed + header
        worksheet_name = row["Worksheet_Name"]
        if worksheet_name:  # Only add link if worksheet exists
            cell = summary_sheet.cell(row=row_idx, column=2)  # Column B
            cell.hyperlink = f"#{worksheet_name}!A1"
            cell.value = worksheet_name
            # Style the hyperlink
            from openpyxl.styles import Font

            cell.font = Font(color="0000FF", underline="single")


def _create_file_worksheet(
    writer: pd.ExcelWriter,
    file_path: Path,
    patterns: list[CredentialPattern],
    worksheet_name: str,
) -> None:
    """Create a worksheet for a single file with its patterns."""
    import pandas as pd

    file_data = []
    for pattern in patterns:
        file_data.append(
            {
                "Pattern_Type": pattern.pattern_type,
                "Risk_Level": pattern.risk_level,
                "Value": pattern.display_value,
                "Line_Start": pattern.line_start,
                "Line_End": pattern.line_end,
                "Column_Start": pattern.column_start,
                "Column_End": pattern.column_end,
                "Context_Before": pattern.context_before,
                "Context_After": pattern.context_after,
                "Confidence": pattern.confidence,
                "Detection_Method": pattern.detection_method,
                "Offset": pattern.offset,
                "Size": pattern.size,
            },
        )

    # Write to worksheet starting at row 3 (0-indexed: startrow=2)
    file_df = pd.DataFrame(file_data)
    file_df.to_excel(writer, sheet_name=worksheet_name, index=False, startrow=2)

    # Add file path in first row if we can access the worksheet
    if hasattr(writer, "book") and hasattr(writer, "sheets"):
        worksheet = writer.sheets[worksheet_name]
        worksheet.cell(row=1, column=1, value=str(file_path))
        # Style the file path cell
        from openpyxl.styles import Font

        worksheet.cell(row=1, column=1).font = Font(bold=True)


def _export_credentials_to_excel(
    result: CredentialAnalysisResult,
    context: ExportContext,
) -> None:
    """Export credential analysis results to Excel with per-file worksheets."""
    import pandas as pd

    from kp_ssf_tools.core.services.excel_export.sheet_management import (
        DefaultSheetNameSanitizer,
    )

    # Generate output filename with timestamp
    timestamp = context.timestamp_service.format_filename_now()
    output_path = Path(f"analyze-credentials-{timestamp}.xlsx")

    # Setup for unique worksheet naming
    sheet_sanitizer = DefaultSheetNameSanitizer()
    used_worksheet_names: set[str] = set()
    worksheet_name_mapping: dict[str, str] = {}  # file_path -> worksheet_name

    # Group patterns by file
    files_with_patterns: dict[Path, list[CredentialPattern]] = {}
    for pattern in result.patterns:
        file_path = pattern.file_path
        if file_path not in files_with_patterns:
            files_with_patterns[file_path] = []
        files_with_patterns[file_path].append(pattern)

    # Determine worksheet names for files with patterns
    for file_path in files_with_patterns:
        worksheet_name = _get_unique_worksheet_name(
            file_path,
            used_worksheet_names,
            sheet_sanitizer,
        )
        worksheet_name_mapping[str(file_path)] = worksheet_name

    try:
        with context.excel_service.workbook_engine.create_writer(output_path) as writer:
            # Create summary worksheet
            summary_data = _build_summary_data(
                result,
                files_with_patterns,
                worksheet_name_mapping,
            )
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

            # Add hyperlinks to summary
            _add_hyperlinks_to_summary(writer, summary_df)

            # Create per-file worksheets only for files that had detected secrets
            for file_path, patterns in files_with_patterns.items():
                worksheet_name = worksheet_name_mapping[str(file_path)]
                _create_file_worksheet(writer, file_path, patterns, worksheet_name)

        context.rich_output.success(f"Results exported to: {output_path}")
    except Exception as e:
        context.rich_output.error(f"Failed to export Excel file: {e}")
        raise
